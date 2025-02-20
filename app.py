from flask import Flask, render_template, request, redirect, url_for, flash, send_file, send_from_directory, make_response, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_migrate import Migrate
from datetime import datetime, date, timedelta
import pytz
import os
import io
import pdfkit
from database import db, CashierReport, User, Settings, Cashier
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell import MergedCell
from sqlalchemy.sql import func

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cashier_check.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key'
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'uploads')

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
migrate = Migrate(app, db)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

with app.app_context():
    db.create_all()

@app.route('/manage_cashiers', methods=['GET', 'POST'])
@login_required
def manage_cashiers():
    if current_user.role != 'admin':
        flash('У вас нет прав для управления кассирами.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        name = request.form['name']
        new_cashier = Cashier(name=name)
        db.session.add(new_cashier)
        db.session.commit()
        flash('Кассир успешно добавлен.', 'success')
    
    cashiers = Cashier.query.all()
    return render_template('manager_cashiers.html', cashiers=cashiers)

@app.route('/delete_cashier/<int:cashier_id>', methods=['POST'])
@login_required
def delete_cashier(cashier_id):
    if current_user.role != 'admin':
        flash('У вас нет прав для удаления кассиров.', 'error')
        return redirect(url_for('index'))

    cashier = Cashier.query.get(cashier_id)
    if cashier:
        db.session.delete(cashier)
        db.session.commit()
        flash('Кассир успешно удален.', 'success')
    else:
        flash('Кассир не найден.', 'error')
    
    return redirect(url_for('manage_cashiers'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
                        flash('Неверные данные. Повторите попытку.')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    difference = None
    cashiers = Cashier.query.all()  # Получаем список кассиров из базы данных

    if request.method == 'POST':
        cashier = request.form['cashier']
        z_report = float(request.form['z_report'])
        humo = float(request.form['humo'])
        uzcard = float(request.form['uzcard'])
        cash = float(request.form['cash'])
        click_payme = float(request.form['click_payme'])
        shift = request.form['shift']
        comments = request.form.get('comments', '')
        reason = request.form.get('reason', '')

        z_report_doc = request.files['z_report_doc']
        humo_receipt = request.files['humo_receipt']
        uzcard_receipt = request.files['uzcard_receipt']
        click_receipts = request.files.getlist('click_receipts')

        debtor_names = request.form.getlist('debtor_name[]')
        debtor_amounts = request.form.getlist('debtor_amount[]')

        total_received = humo + uzcard + cash + click_payme
        difference = total_received - z_report

        z_report_doc_path = secure_filename(z_report_doc.filename)
        z_report_doc.save(os.path.join(app.config['UPLOAD_FOLDER'], z_report_doc_path))

        humo_receipt_path = secure_filename(humo_receipt.filename)
        humo_receipt.save(os.path.join(app.config['UPLOAD_FOLDER'], humo_receipt_path))

        uzcard_receipt_path = secure_filename(uzcard_receipt.filename)
        uzcard_receipt.save(os.path.join(app.config['UPLOAD_FOLDER'], uzcard_receipt_path))

        click_receipts_paths = []
        for receipt in click_receipts:
            receipt_path = secure_filename(receipt.filename)
            receipt.save(os.path.join(app.config['UPLOAD_FOLDER'], receipt_path))
            click_receipts_paths.append(receipt_path)

        settings = Settings.query.first()
        timezone = pytz.timezone(settings.timezone) if settings else pytz.timezone('Asia/Tashkent')
        current_time = datetime.now(timezone)

        report = CashierReport(
            cashier=cashier,
            z_report=z_report,
            humo=humo,
            uzcard=uzcard,
            cash=cash,
            click_payme=click_payme,
            difference=difference,
            shift=shift,
            comments=comments,
            reason=reason,
            timestamp=current_time,
            z_report_doc=z_report_doc_path,
            humo_receipt=humo_receipt_path,
            uzcard_receipt=uzcard_receipt_path,
            click_receipts=";".join(click_receipts_paths),
            debtor_names=";".join(debtor_names),
            debtor_amounts=";".join(debtor_amounts)
        )
        db.session.add(report)
        db.session.commit()

    settings = Settings.query.first()
    return render_template('index.html', difference=difference, settings=settings, cashiers=cashiers)

from datetime import datetime
from flask import request, flash

@app.route('/reports', methods=['GET'])
@login_required
def reports():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    cashier = request.args.get('cashier')
    search = request.args.get('search')

    query = CashierReport.query

    # Проверяем формат даты и фильтруем
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
            query = query.filter(CashierReport.timestamp >= start_date_obj)
        except ValueError:
            flash("Ошибка: Неверный формат даты", "error")

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(CashierReport.timestamp <= end_date_obj)
        except ValueError:
            flash("Ошибка: Неверный формат даты", "error")

    # Фильтр по кассиру
    if cashier:
        query = query.filter(CashierReport.cashier == cashier)

    # Фильтр по поиску
    if search:
        query = query.filter(
            (CashierReport.cashier.ilike(f'%{search}%')) |
            (CashierReport.shift.ilike(f'%{search}%')) |
            (CashierReport.comments.ilike(f'%{search}%'))
        )

    reports = query.all()
    settings = Settings.query.first()
    timezone = pytz.timezone(settings.timezone) if settings else pytz.timezone('Asia/Tashkent')

    # Проверяем, есть ли данные
    if not reports:
        flash("⚠ Отчеты не найдены за выбранный период.", "warning")

    # Применяем часовой пояс
    for report in reports:
        report.timestamp = report.timestamp.astimezone(timezone)

    return render_template('reports.html', reports=reports, cashiers=Cashier.query.all(), settings=settings)


from datetime import datetime, timedelta
from flask import request

@app.route('/summary', methods=['GET'])
@login_required
def summary():
    # Получаем текущую дату
    today = datetime.today().strftime('%Y-%m-%d')

    # Получаем параметры запроса (если пользователь выбрал другую дату)
    start_date = request.args.get('start_date', today)  # По умолчанию = сегодня
    end_date = request.args.get('end_date', today)  # По умолчанию = сегодня

    query = CashierReport.query

    # Фильтр по дате (преобразуем в datetime)
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)

    query = query.filter(CashierReport.timestamp >= start_date_obj,
                         CashierReport.timestamp <= end_date_obj)

    reports = query.all()

    # Загружаем настройки
    settings = Settings.query.first()
    if not settings:
        settings = type('Settings', (), {"currency": "сум"})()

    # Считаем итоги
    total_z_report = sum([r.z_report for r in reports])
    total_humo = sum([r.humo for r in reports])
    total_uzcard = sum([r.uzcard for r in reports])
    total_cash = sum([r.cash for r in reports])
    total_click_payme = sum([r.click_payme for r in reports])
    total_difference = sum([r.difference for r in reports])

    # Генерация данных для графика
    graph_data = {
        "data": [
            {
                "x": [r.timestamp.strftime('%d.%m.%Y') for r in reports],
                "y": [r.z_report + r.humo + r.uzcard + r.cash + r.click_payme for r in reports],
                "type": "bar",
                "marker": {"color": "blue"}
            }
        ],
        "layout": {
            "title": "Выручка по дням",
            "xaxis": {"title": "Дата"},
            "yaxis": {"title": "Сумма"}
        }
    }

    return render_template('summary.html', reports=reports, settings=settings, 
                           total_z_report=total_z_report, total_humo=total_humo, 
                           total_uzcard=total_uzcard, total_cash=total_cash, 
                           total_click_payme=total_click_payme, 
                           total_difference=total_difference, 
                           graph_json=graph_data, 
                           selected_start_date=start_date, selected_end_date=end_date)


@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    settings = Settings.query.first()
    if request.method == 'POST':
        if not settings:
            settings = Settings()
        settings.currency = request.form['currency'] or 'сум'
        settings.language = request.form['language'] or 'ru'
        settings.timezone = request.form['timezone'] or 'Asia/Tashkent'
        db.session.add(settings)
        db.session.commit()
        flash('Настройки успешно сохранены.', 'success')
    return render_template('settings.html', settings=settings)

@app.route('/delete_report/<int:report_id>', methods=['POST'])
@login_required
def delete_report(report_id):
    if current_user.role == 'admin':
        report = CashierReport.query.get(report_id)
        if report:
            db.session.delete(report)
            db.session.commit()
            flash('Отчет успешно удален.', 'success')
        else:
            flash('Отчет не найден.', 'error')
    else:
        flash('У вас нет прав на удаление отчетов.', 'error')
    return redirect(url_for('reports'))

@app.route('/edit_report/<int:report_id>', methods=['GET', 'POST'])
@login_required
def edit_report(report_id):
    report = CashierReport.query.get(report_id)
    if not report:
        flash('Отчет не найден.', 'error')
        return redirect(url_for('reports'))

    if request.method == 'POST':
        if current_user.role == 'manager' or current_user.role == 'admin':
            report.z_report = float(request.form['z_report'])
            report.humo = float(request.form['humo'])
            report.uzcard = float(request.form['uzcard'])
            report.cash = float(request.form['cash'])
            report.click_payme = float(request.form['click_payme'])
            report.reason = request.form.get('reason', '')

            debtor_names = request.form.getlist('debtor_name[]')
            debtor_amounts = request.form.getlist('debtor_amount[]')
            report.debtor_names = ";".join(debtor_names)
            report.debtor_amounts = ";".join(debtor_amounts)

            total_received = report.humo + report.uzcard + report.cash + report.click_payme
            report.difference = total_received - report.z_report

            db.session.commit()
            flash('Отчет успешно обновлен.', 'success')
        else:
            flash('У вас нет прав на редактирование отчетов.', 'error')
        return redirect(url_for('reports'))
    
    return render_template('edit_report.html', report=report, settings=Settings.query.first(), zip=zip)


@app.route('/download_excel/<int:report_id>')
@login_required
def download_excel(report_id):
    report = CashierReport.query.get(report_id)
    settings = Settings.query.first()

    if not report:
        flash('Отчет не найден.', 'error')
        return redirect(url_for('reports'))

    if not settings:
        class DefaultSettings:
            currency = 'сум'
        settings = DefaultSettings()

    # Создание Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет сверки"

    # Настройка стилей
    header_font = Font(bold=True, color="FFFFFF", size=14)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    data_font = Font(size=12)
    title_font = Font(size=20, bold=True, color="4F81BD")
    alignment_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовок
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "Отчет о сдаче выручки"
    title_cell.font = title_font
    title_cell.alignment = alignment_center

        # Логотип (если есть логотип, добавьте его. В противном случае удалите этот блок)
    logo_path = 'path/to/logo.png'  # Путь к вашему логотипу
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.height = 60  # Устанавливаем высоту изображения
        logo.width = 120  # Устанавливаем ширину изображения
        ws.add_image(logo, "E1")

    # Информация об отчете
    data = [
        ["Отчет №", report.id, "", ""],
        ["Кассир", report.cashier, "", ""],
        ["Смена", report.shift, "", ""],
        [f"Сумма по Z-отчёту ({settings.currency})", report.z_report, "", ""],
        [f"HUMO ({settings.currency})", report.humo, "", ""],
        [f"UZCARD ({settings.currency})", report.uzcard, "", ""],
        [f"Наличные ({settings.currency})", report.cash, "", ""],
        [f"Click/Payme ({settings.currency})", report.click_payme, "", ""],
        [f"Разница ({settings.currency})", report.difference, "", ""],
        ["Комментарии", report.comments, "", ""],
        ["Причина", report.reason, "", ""],
        ["Дата", report.timestamp.strftime('%d.%m.%Y %H:%M'), "", ""]
    ]

    # Заполнение данных
    for row in data:
        ws.append(row)

    # Форматирование заголовков и данных
    for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.column_letter == "A":
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
            else:
                cell.font = data_font

    # Настройка ширины колонок
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col if not isinstance(cell, MergedCell)]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if column:  # Проверка на пустую колонку
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Сохранение Excel файла во временное хранилище
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name=f'report_{report.id}.xlsx', as_attachment=True)

import io
from flask import send_file

@app.route('/download_all_excel')
@login_required
def download_all_excel():
    reports = CashierReport.query.all()
    settings = Settings.query.first()

    # Создание Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Все отчеты кассиров"

    # Настройка стилей
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    data_font = Font(size=11)
    bold_font = Font(bold=True, color="1F4E79")
    total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Заголовки таблицы
    headers = ["ID", "Кассир", "Смена", "Z-отчёт", "HUMO", "UZCARD", "Наличные", "Click/Payme", "Разница", "Комментарии", "Причина", "Дата"]
    ws.append(headers)

    for col in ws.columns:
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border

    # Заполнение данными
    total_z_report = 0
    total_humo = 0
    total_uzcard = 0
    total_cash = 0
    total_click_payme = 0
    total_difference = 0

    for report in reports:
        data = [
            report.id,
            report.cashier,
            report.shift,
            report.z_report,
            report.humo,
            report.uzcard,
            report.cash,
            report.click_payme,
            report.difference,
            report.comments,
            report.reason,
            report.timestamp.strftime('%d.%m.%Y %H:%M')
        ]
        ws.append(data)

        total_z_report += report.z_report
        total_humo += report.humo
        total_uzcard += report.uzcard
        total_cash += report.cash
        total_click_payme += report.click_payme
        total_difference += report.difference

    # Форматирование данных
    for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = alignment_left
            cell.border = thin_border

    # Итоговые суммы под каждой категорией
    ws.append([""])
    ws.append(["Общие суммы"])
    ws.append(["Общая сумма по Z-отчету:", "", "", total_z_report])
    ws.append(["Общая сумма HUMO:", "", "", total_humo])
    ws.append(["Общая сумма UZCARD:", "", "", total_uzcard])
    ws.append(["Общая сумма наличных:", "", "", total_cash])
    ws.append(["Общая сумма по Click/Payme:", "", "", total_click_payme])
    ws.append(["Общая разница:", "", "", total_difference])

    # Объединение ячеек и стиль итоговых строк
    for row in ws.iter_rows(min_row=ws.max_row - 6, max_col=4, max_row=ws.max_row):
        for cell in row:
            cell.font = bold_font
            cell.fill = total_fill
            cell.alignment = alignment_center
            cell.border = thin_border
        ws.merge_cells(start_row=row[0].row, start_column=2, end_row=row[0].row, end_column=3)

    # Настройка ширины колонок
    for col in ws.columns:
        max_length = 0
        column = [cell for cell in col if not isinstance(cell, MergedCell)]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Сохранение Excel файла во временное хранилище
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='all_cashier_reports.xlsx', as_attachment=True)


import plotly.graph_objs as go
from flask import render_template

@app.route('/charts')
@login_required
def charts():
    reports = CashierReport.query.all()
    settings = Settings.query.first()

    dates = [report.timestamp.strftime('%Y-%m-%d') for report in reports]
    z_reports = [report.z_report for report in reports]
    cash = [report.cash for report in reports]
    humo = [report.humo for report in reports]
    uzcard = [report.uzcard for report in reports]
    click_payme = [report.click_payme for report in reports]

    fig = go.Figure()
    fig.add_trace(go.Bar(x=dates, y=z_reports, name='Z-отчёт'))
    fig.add_trace(go.Bar(x=dates, y=cash, name='Наличные'))
    fig.add_trace(go.Bar(x=dates, y=humo, name='HUMO'))
    fig.add_trace(go.Bar(x=dates, y=uzcard, name='UZCARD'))
    fig.add_trace(go.Bar(x=dates, y=click_payme, name='Click/Payme'))

    fig.update_layout(
        barmode='group', 
        title='Сравнение сумм по дням', 
        xaxis_title='Дата', 
        yaxis_title='Сумма',
        template='plotly_white'
    )

    graph_json = fig.to_json()

    return render_template('charts.html', graph_json=graph_json)


@app.route('/generate_report_by_period/<period>')
def generate_report_by_period(period):
    if period == "day":
        date_filter = datetime.now()
    elif period == "week":
        date_filter = datetime.now() - timedelta(weeks=1)
    elif period == "month":
        date_filter = datetime.now() - timedelta(days=30)
    else:
        return "Некорректный период.", 400

    reports = CashierReport.query.filter(CashierReport.timestamp >= date_filter).all()
    if not reports:
        return f"Нет отчетов за выбранный период: {period}.", 200

    report_text = f"📊 *Отчет за {period}:*\n\n"
    for report in reports:
        report_text += (
            f"👤 *Кассир:* {report.cashier}\n"
            f"💵 *Сумма по Z-отчету:* {report.z_report}\n"
            f"💳 *HUMO:* {report.humo}\n"
            f"💳 *UZCARD:* {report.uzcard}\n"
            f"💸 *Наличные:* {report.cash}\n"
            f"📱 *Click/Payme:* {report.click_payme}\n"
            f"📉 *Разница:* {report.difference}\n"
            f"📝 *Комментарии:* {report.comments or 'Нет комментариев'}\n"
            f"📅 *Дата:* {report.timestamp}\n\n"
            "----------------------------------------\n\n"
        )
    return report_text

@app.route('/generate_total_report', methods=['GET'])
def generate_custom_total_report():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    if not start_date_str or not end_date_str:
        return "Ошибка: Укажите даты (start_date, end_date).", 400

    try:
        timezone = pytz.timezone("Asia/Tashkent")
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').replace(hour=0, minute=0, second=0)
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        return "Ошибка: Некорректный формат даты. Используйте YYYY-MM-DD.", 400

    return generate_total_report_for_dates(start_date, end_date)

def generate_total_report_for_dates(start_date, end_date):
    with app.app_context():
        total_z_report = db.session.query(db.func.sum(CashierReport.z_report)).filter(
            CashierReport.timestamp >= start_date,
            CashierReport.timestamp <= end_date
        ).scalar() or 0

    report_text = (
        f"📊 *Итоговый отчёт:*\n"
        f"📅 *Период:* {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}\n\n"
        f"💵 *Сумма по Z-отчету:* {total_z_report}\n"
    )

    return report_text



from flask import render_template, redirect, url_for, request
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from sqlalchemy.sql import func
from database import db, CashierReport, User, Settings, Cashier  # ✅ Исправленный импорт

@app.route('/dashboard')
@login_required
def dashboard():
    # Проверяем, админ ли это
    if current_user.username != "admin":
        return redirect(url_for('index'))  # Если не админ, отправляем на главную

    # Считаем общую сумму денег в кассах
    total_z_report = db.session.query(func.sum(CashierReport.z_report)).scalar() or 0
    total_humo = db.session.query(func.sum(CashierReport.humo)).scalar() or 0
    total_uzcard = db.session.query(func.sum(CashierReport.uzcard)).scalar() or 0
    total_cash = db.session.query(func.sum(CashierReport.cash)).scalar() or 0
    total_click_payme = db.session.query(func.sum(CashierReport.click_payme)).scalar() or 0

    # Количество проблемных отчетов (где разница не 0)
    problem_reports_count = CashierReport.query.filter(CashierReport.difference != 0).count()

    # Пагинация (разбиение проблемных отчетов на страницы)
    page = request.args.get('page', 1, type=int)  # Получаем текущую страницу (по умолчанию 1)
    per_page = 10  # Количество отчетов на одной странице

    # Получаем проблемные отчеты (все, но с разбиением по страницам)
    problem_reports = (
        CashierReport.query.filter(CashierReport.difference != 0)
        .order_by(CashierReport.timestamp.desc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    # Лучший и худший кассир (по количеству отчетов)
    best_cashier = (
        db.session.query(CashierReport.cashier, func.count(CashierReport.id))
        .group_by(CashierReport.cashier)
        .order_by(func.count(CashierReport.id).desc())
        .first()
    )
    worst_cashier = (
        db.session.query(CashierReport.cashier, func.count(CashierReport.id))
        .group_by(CashierReport.cashier)
        .order_by(func.count(CashierReport.id).asc())
        .first()
    )

    # Выручка за последние 7 дней (для графика)
    last_7_days = datetime.utcnow() - timedelta(days=7)
    sales_data = (
        db.session.query(func.date(CashierReport.timestamp), func.sum(CashierReport.z_report))
        .filter(CashierReport.timestamp >= last_7_days)
        .group_by(func.date(CashierReport.timestamp))
        .all()
    )

    graph_data = {
        "labels": [str(date) for date, _ in sales_data],
        "values": [total for _, total in sales_data],
    }

    return render_template(
        "dashboard.html",
        total_z_report=total_z_report,
        total_humo=total_humo,
        total_uzcard=total_uzcard,
        total_cash=total_cash,
        total_click_payme=total_click_payme,
        problem_reports_count=problem_reports_count,
        problem_reports=problem_reports,  # Теперь передаем объект с пагинацией!
        best_cashier=best_cashier,
        worst_cashier=worst_cashier,
        graph_data=graph_data,
    )

if __name__ == '__main__':
    app.run(debug=True)

